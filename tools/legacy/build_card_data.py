"""
Build the single source-of-truth spreadsheet from the CURRENT design doc, Image Cards.docx.
(Image Cards.docx is more up to date than Cards.docx.)

In this doc every card field is a separate, internally-unordered text box, but a card's
boxes are contiguous in document order. We greedily accumulate one of each field type and
flush a card when a field would be overwritten by a different value or a new name appears.

Re-run after editing the doc to regenerate Cards_Data.xlsx.
Columns: Name | Category | Tier/Subtype | Class | Mana Cost | Effect/Text | Image File | Copies | Notes
"""
import zipfile, re, os, difflib
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DOCX = os.path.join(_ROOT, "assets", "Image Cards.docx")
OUT = os.path.join(_ROOT, "data", "Cards_Data.xlsx")
_IMAGES = os.path.join(_ROOT, "assets", "Images")

# -------------------------------------------------------------- text extraction
def paragraphs(docx):
    xml = zipfile.ZipFile(docx).read("word/document.xml").decode("utf-8", "ignore")
    xml = xml.replace("</w:p>", "\n###P###\n").replace("<w:br/>", "\n")
    txt = re.sub(r"<[^>]+>", "", xml).replace("&amp;", "&").replace("’", "'")
    return [p.strip("\n") for p in txt.split("###P###") if p.strip()]

# category label -> (Category, Tier)
TIERMAP = {
    "Mana": ("Mana", ""), "Moderate Mana": ("Mana", "Moderate"), "Greater Mana": ("Mana", "Greater"),
    "Support": ("Support", ""),
    "Light Weapon": ("Weapon", "Light"), "Heavy Weapon": ("Weapon", "Heavy"),
    "Common Artifacts": ("Artifact", "Common"), "Ancient Artifacts": ("Artifact", "Ancient"),
    "Minion": ("Minion", ""), "Disaster": ("Disaster", ""),
}
CLASSES = {"Blacksmith", "Wizard", "Paladin", "Cleric", "Enchanter", "Ranger",
           "Bard", "Druid", "Weaponmaster"}
SKIP_NAMES = {"Turn Overview", "Turn Marker", "BOSS", "JOKER", "Contents:", "Regeln:", ""}

def classify(p):
    name_m = re.match(r"^(?:\d{5,}|00)\n\s*\n\s*(.+)$", p, re.S)
    if name_m:
        return ("name", re.sub(r"\s+", " ", name_m.group(1)).strip())
    r = re.sub(r"^(?:\d{5,}|00)\s*", "", p).strip()
    r1 = re.sub(r"\s+", " ", r)
    if not r1:
        return ("anchor", None)
    if r1.startswith("Effect:"):
        return ("effect", re.sub(r"^Effect:\s*", "", r1).strip())
    if re.fullmatch(r"\d{1,2}", r1):
        return ("cost", int(r1))
    if "Red Dragon" in r1:               # boss/minion/disaster footer
        return ("boss_footer", None)
    if re.fullmatch(r"(Weapon|Artifact|Mana) - [A-Za-z]+", r1):  # class starter
        return ("clscat", r1)
    if r1 in TIERMAP:
        return ("category", r1)
    if r1 in CLASSES:
        return ("charclass", r1)
    return ("other", r1)

# -------------------------------------------------------------- parse
def parse(docx):
    cards = []
    cur = {}

    def flush():
        nonlocal cur
        if cur.get("name") and cur["name"] not in SKIP_NAMES \
                and not re.fullmatch(r"[\d /]+", cur["name"]) \
                and not cur["name"].startswith("Player "):
            cards.append(cur)
        cur = {}

    for p in paragraphs(docx):
        t, v = classify(p)
        if t == "name":
            if cur.get("name") == v:
                continue                 # '00' duplicate of the same name box
            if cur.get("name"):
                flush()
            cur["name"] = v
        elif t == "effect":
            if cur.get("effect") and cur["effect"] != v:
                flush()
            cur["effect"] = v
        elif t in ("category", "clscat", "charclass"):
            if cur.get("cat_t") and cur["cat_t"] != t:
                pass
            if cur.get("category") and cur["category"] != v:
                flush()
            cur["category"] = v
            cur["cat_t"] = t
        elif t == "cost":
            cur["cost"] = v
        elif t == "boss_footer":
            cur["boss"] = True
        # 'other'/'anchor' ignored (rules text, HP/Charge tracks, page numbers)
    flush()
    return cards

# -------------------------------------------------------------- normalize a card
def normalize(c):
    name = c["name"]
    effect = c.get("effect", "")
    cost = c.get("cost", "")
    cat_t = c.get("cat_t")
    raw = c.get("category", "")
    klass = "All"
    category, tier = "", ""
    if cat_t == "charclass":            # character card
        category, klass, cost = "Character", raw, ""
    elif cat_t == "clscat":             # class starter (e.g. "Weapon - Blacksmith")
        base, kl = [x.strip() for x in raw.split("-", 1)]
        category, klass = base, kl
    elif raw in TIERMAP:
        category, tier = TIERMAP[raw]
    elif c.get("boss"):
        category = "Boss"
    if c.get("boss") and category not in ("Minion", "Disaster"):
        category = category or "Boss"
        if category not in ("Minion", "Disaster"):
            category = "Boss"
    return dict(name=name, category=category or "?", tier=tier, klass=klass,
                cost=cost, effect=effect)

# -------------------------------------------------------------- image matching
def norm(s):
    return re.sub(r"[^a-z0-9]", "", s.lower())

ALIAS = {
    "Magic Crystal": [
        "Images/Characters/Magic Crystal/pavoras_blue_Magical_Mana_crystals_floating_in_a_fantasy_settin_033423e9-4372-4daa-a844-2339dee50ad7.png",
        "Images/Characters/Magic Crystal/pavoras_blue_Magical_Mana_crystals_floating_in_a_fantasy_settin_34346a12-5c01-4fe9-b3c8-434e08ef5138.png",
    ],
    "Wandering Wisp": ["Images/artifacts/wisp1.png", "Images/artifacts/wisp2.png"],
    "Multiplier Maul": ["Images/weapons/maul of multiplication.png"],
    "Groveguard Bow": ["Images/weapons/groveguard knuckles.png"],
}

def build_image_index():
    items = []
    for root, _, files in os.walk(_IMAGES):
        for f in files:
            if f.lower().endswith(".png"):
                rel = os.path.join(root, f).replace(os.sep, "/")
                base = re.sub(r"\s*\d+$", "", os.path.splitext(f)[0]).strip()
                items.append((norm(base), base, rel))
    return items

def match_image(name, items):
    if name in ALIAS:
        return ";".join(ALIAS[name]), "manual alias - reconcile name vs art filename"
    n = norm(name)
    exact = sorted({rel for nb, db, rel in items if nb == n})
    if exact:
        return ";".join(exact), ""
    sub = sorted({rel for nb, db, rel in items
                  if nb and len(min(nb, n, key=len)) >= 5 and (nb in n or n in nb)})
    if sub:
        return ";".join(sub), "image filename differs - verify"
    nmap = defaultdict(list)
    for nb, db, rel in items:
        nmap[nb].append((db, rel))
    close = difflib.get_close_matches(n, list(nmap), n=1, cutoff=0.72)
    if close:
        db = nmap[close[0]][0][0]
        return ";".join(sorted({rel for _, rel in nmap[close[0]]})), f"image filename differs ('{db}') - verify"
    guess = difflib.get_close_matches(n, list(nmap), n=1, cutoff=0.45)
    hint = f" (closest: '{nmap[guess[0]][0][0]}')" if guess else ""
    return "", f"NO IMAGE FOUND{hint}"

# -------------------------------------------------------------- main
if __name__ == "__main__":
    raw_cards = [normalize(c) for c in parse(DOCX)]

    groups = defaultdict(list)
    for c in raw_cards:
        groups[c["name"]].append(c)

    # collapse phantom fragments: a '?'-category variant whose effect duplicates a
    # properly-categorized variant of the same name is a parse artifact -> drop it.
    for name, variants in groups.items():
        real = [v for v in variants if v["category"] != "?"]
        real_effs = {re.sub(r"\s+", " ", v["effect"]).strip().lower() for v in real}
        if real:
            groups[name] = real + [v for v in variants if v["category"] == "?"
                                   and re.sub(r"\s+", " ", v["effect"]).strip().lower() not in real_effs]

    img_idx = build_image_index()
    rows = []
    for name, variants in groups.items():
        # one row per card: pick a primary variant, flag anything that differs to verify.
        pool = [v for v in variants if v["category"] != "?"] or variants
        primary = sorted(pool, key=lambda v: (v["cost"] == "",))[0]  # prefer one with a cost
        costs = sorted({str(v["cost"]) for v in variants if v["cost"] != ""})
        cats = sorted({f"{v['category']}/{v['tier']}".rstrip("/") for v in variants if v["category"] != "?"})
        effs = {re.sub(r"\s+", " ", v["effect"]).strip() for v in variants if v["effect"]}
        img, imgnote = match_image(name, img_idx)
        notes = []
        if len(costs) > 1:
            notes.append(f"verify cost (extraction saw: {', '.join(costs)})")
        if len(cats) > 1:
            notes.append(f"verify category/grouping (saw: {', '.join(cats)})")
        if len(effs) > 1:
            notes.append("verify effect (multiple texts extracted - possible field bleed)")
        if primary["category"] == "?":
            notes.append("category not captured - fill in")
        if imgnote:
            notes.append(imgnote)
        rows.append(dict(name=name, category=primary["category"], tier=primary["tier"],
                         klass=primary["klass"], cost=primary["cost"], effect=primary["effect"],
                         image=img, copies="", notes="; ".join(notes)))

    # Village card (from locked design decisions, not in source doc)
    rows.append(dict(
        name="Village", category="Village", tier="", klass="", cost="",
        effect="HP 40. Has its own Charge track. Ability: deal 2 DMG. "
               "Ultimate (3 Charge): deal 6 DMG. Triggered by a player's Use Ability (3M) or by cards.",
        image=";".join(f"Images/Characters/Village/{f}" for f in
                       ["Village.png", "Village1.png", "Village2.png", "Village3.png", "Village4.png"]),
        copies=1, notes="added from locked design decisions (not in source doc)"))

    used = {p for r in rows for p in (r["image"] or "").split(";") if p}
    orphans = [rel for _, _, rel in img_idx if rel not in used]

    order = {c: i for i, c in enumerate(
        ["Character", "Village", "Mana", "Weapon", "Artifact", "Support", "Boss", "Minion", "Disaster"])}
    rows.sort(key=lambda r: (order.get(r["category"], 99), r["tier"], r["name"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cards"
    ws.append(["Name", "Category", "Tier/Subtype", "Class", "Mana Cost", "Effect/Text",
               "Image File", "Copies", "Notes/Conflicts"])
    hfill = PatternFill("solid", fgColor="2F5496")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
    warn = PatternFill("solid", fgColor="FFE699")
    for r in rows:
        ws.append([r["name"], r["category"], r["tier"], r["klass"], r["cost"],
                   r["effect"], r["image"], r["copies"], r["notes"]])
        if r["notes"]:
            for c in ws[ws.max_row]:
                c.fill = warn
    for i, w in enumerate([22, 13, 12, 13, 10, 62, 40, 8, 34], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Orphan Images")
    ws2.append(["Images with no matched card"])
    ws2["A1"].font = Font(bold=True)
    for o in sorted(set(orphans)):
        ws2.append([o])
    ws2.column_dimensions["A"].width = 70
    wb.save(OUT)

    print("Wrote", OUT)
    print("rows:", len(rows), "| unique names:", len(groups), "| orphans:", len(set(orphans)))
    print("by category:", dict(Counter(r["category"] for r in rows)))
    print("conflicts:", sum(1 for r in rows if "CONFLICT" in r["notes"]))
    print("no-image:", sum(1 for r in rows if "NO IMAGE" in r["notes"]))
    print("unknown-category:", sum(1 for r in rows if r["category"] == "?"))
