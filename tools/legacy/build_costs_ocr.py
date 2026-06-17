"""
Authoritative costs via the PDF cost TEXT layer (exact), with OCR only to identify which
card each cell holds. Two stages, with the slow OCR cached to ocr_cache.json:
  Stage 1 (slow, once): read each cell's exact cost from PDF text + OCR its effect -> cache.
  Stage 2 (fast): fuzzy-match cached effect to the docx card list, majority-vote cost, merge.
Re-run after deleting ocr_cache.json to re-OCR; otherwise it only re-does the fast merge.
Writes Cards_Data.xlsx (docx text + PDF-voted cost).
"""
import re, io, json, os, difflib
from collections import defaultdict, Counter
import fitz, pytesseract
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import build_card_data as B

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PDF = os.path.join(_ROOT, "assets", "Image Cards.pdf")
OUT = os.path.join(_ROOT, "data", "Cards_Data.xlsx")
CACHE = os.path.join(_ROOT, "data", "caches", "ocr_cache.json")
NO_COST = {"Boss", "Minion", "Disaster", "Character", "Village"}

def norme(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())

# ---- Stage 1: OCR cache (cost + effect text per filled cell) -------------------
COLX = [198, 391, 582, 772]
def cell_of(x, y):
    return (0 if y < 180 else 1, min(range(4), key=lambda i: abs(x - COLX[i])))

if os.path.exists(CACHE):
    cells = json.load(open(CACHE))
    print("loaded OCR cache:", len(cells), "cells")
else:
    doc = fitz.open(PDF)
    cells = []
    for pidx in range(doc.page_count):
        pg = doc[pidx]
        costmap = {}
        for w in pg.get_text("words"):
            if re.fullmatch(r"\d{1,2}", w[4]):
                costmap[cell_of(w[0], w[1])] = int(w[4])
        if not costmap:
            continue
        img = Image.open(io.BytesIO(pg.get_pixmap(matrix=fitz.Matrix(3, 3)).tobytes("png")))
        cw, ch = img.width / 4, img.height / 2
        for (r, c), cost in costmap.items():
            cell = img.crop((int(c*cw), int(r*ch), int((c+1)*cw), int((r+1)*ch)))
            lines = [l.strip() for l in pytesseract.image_to_string(cell, config="--psm 6").splitlines() if l.strip()]
            ei = next((i for i, l in enumerate(lines) if "ffect" in l.lower()), None)
            eff = re.sub(r"^.*?ffect[:\s]*", "", " ".join(lines[ei:ei+4])) if ei is not None else ""
            cells.append([cost, eff])
        print(f"page {pidx} ok ({len(costmap)} cells)")
    json.dump(cells, open(CACHE, "w"))
    print("wrote OCR cache:", len(cells), "cells")

# ---- known cards from docx -----------------------------------------------------
raw = [B.normalize(c) for c in B.parse(B.DOCX)]
groups = defaultdict(list)
for c in raw:
    groups[c["name"]].append(c)
known = []
for name, vs in groups.items():
    pool = [v for v in vs if v["category"] != "?"] or vs
    p = sorted(pool, key=lambda v: (v["cost"] == "",))[0]
    cats = sorted({f"{v['category']}/{v['tier']}".rstrip("/") for v in vs if v["category"] != "?"})
    known.append(dict(name=name, category=p["category"], tier=p["tier"], klass=p["klass"],
                      effect=p["effect"], cats_seen=cats))
known_norm = [(k, norme(k["effect"])) for k in known if len(norme(k["effect"])) >= 4]

# ---- Stage 2: match + vote (keep match confidence) -----------------------------
votes = defaultdict(list)          # name -> list of (cost, ratio)
for cost, eff in cells:
    en = norme(eff)
    if len(en) < 5:
        continue
    best, br = None, 0.0
    for k, kn in known_norm:
        rt = difflib.SequenceMatcher(None, en, kn).ratio()
        if rt > br:
            best, br = k, rt
    if best and br >= 0.70:         # stricter: fewer mis-assignments
        votes[best["name"]].append((cost, br))

# ---- merge ---------------------------------------------------------------------
img_idx = B.build_image_index()
rows = []
for k in known:
    name, notes = k["name"], []
    v = votes.get(name)
    cost = ""
    if v:
        tally = Counter(c for c, _ in v)
        ranked = tally.most_common()
        cost, top = ranked[0]
        conf = max(r for c, r in v if c == cost)
        if len(ranked) > 1 and ranked[1][1] >= top:
            notes.append(f"verify cost (OCR split: {dict(tally)})")
        elif conf < 0.80:
            notes.append(f"verify cost (low match confidence {conf:.2f})")
    elif norme(k["effect"]) in ("deal1dmg",) and k["category"] in ("Weapon", "Artifact"):
        cost = 0                                   # class-starter basic attack
    elif k["category"] not in NO_COST:
        notes.append("cost: no OCR match - fill in")
    if len(k["cats_seen"]) > 1:
        notes.append(f"verify category (saw: {', '.join(k['cats_seen'])})")
    if k["category"] == "?":
        notes.append("category not captured - fill in")
    img, imgnote = B.match_image(name, img_idx)
    if imgnote:
        notes.append(imgnote)
    rows.append(dict(name=name, category=k["category"], tier=k["tier"], klass=k["klass"],
                     cost=cost, effect=k["effect"], image=img, notes="; ".join(notes)))

rows.append(dict(name="Village", category="Village", tier="", klass="", cost="",
    effect="HP 40. Has its own Charge track. Ability: deal 2 DMG. Ultimate (3 Charge): deal 6 DMG. "
           "Triggered by a player's Use Ability (3M) or by cards.",
    image=";".join(f"Images/Characters/Village/{f}" for f in
                   ["Village.png", "Village1.png", "Village2.png", "Village3.png", "Village4.png"]),
    notes=""))

order = {c: i for i, c in enumerate(
    ["Character", "Village", "Mana", "Weapon", "Artifact", "Support", "Boss", "Minion", "Disaster"])}
rows.sort(key=lambda r: (order.get(r["category"], 99), str(r["tier"]), r["name"]))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Cards"
ws.append(["Name", "Category", "Tier/Subtype", "Class", "Mana Cost", "Effect/Text", "Image File", "Notes/Conflicts"])
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="2F5496")
warn = PatternFill("solid", fgColor="FFE699")
for r in rows:
    ws.append([r["name"], r["category"], r["tier"], r["klass"], r["cost"], r["effect"], r["image"], r["notes"]])
    if r["notes"]:
        for c in ws[ws.max_row]:
            c.fill = warn
for i, w in enumerate([22, 13, 12, 13, 10, 62, 40, 38], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"
wb.save(OUT)

supply = {"Mana", "Weapon", "Artifact", "Support"}
print("Wrote", OUT, "| cards:", len(rows))
print("supply w/ cost:", sum(1 for r in rows if r["category"] in supply and r["cost"] != ""),
      "/", sum(1 for r in rows if r["category"] in supply))
print("flagged:", sum(1 for r in rows if r["notes"]))
