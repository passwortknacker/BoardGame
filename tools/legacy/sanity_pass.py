"""
Operate IN PLACE on the user-edited Cards_Data.xlsx:
  1. Rename image files to match card names; update the Image File column.
  2. Clear Mana Cost on Boss / Minion / Disaster cards (they have no cost).
  3. Add a 'Sanity Check' column flagging cards whose effect/category/cost look wrong.
Does NOT touch the user's Notes column or any other data.
"""
import openpyxl, os, re, shutil
from openpyxl.styles import Font, Alignment, PatternFill

XLSX = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                    "data", "Cards_Data.xlsx")
wb = openpyxl.load_workbook(XLSX)
ws = wb["Cards"]

# header indices (1-based): 1 Name 2 Cat 3 Tier 4 Class 5 Cost 6 Effect 7 Image 8 Notes
SANCOL = 9
ws.cell(row=1, column=SANCOL, value="Sanity Check")
ws.cell(row=1, column=SANCOL).font = Font(bold=True, color="FFFFFF")
ws.cell(row=1, column=SANCOL).fill = PatternFill("solid", fgColor="C00000")
ws.column_dimensions[openpyxl.utils.get_column_letter(SANCOL)].width = 60

INVALID = '<>:"/\\|?*'
def safe(s):
    return "".join(c for c in s if c not in INVALID).strip()

# ---- 1. rename image files to match card names --------------------------------
renames = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 1).value
    img = ws.cell(r, 7).value
    if not name or not img:
        continue
    files = [f for f in img.split(";") if f.strip()]
    multi = len(files) > 1
    newpaths = []
    for i, f in enumerate(files):
        f = f.strip()
        if not os.path.exists(f):
            newpaths.append(f)
            continue
        folder = os.path.dirname(f)
        base = safe(name) + (f" {i+1}" if multi else "")
        newpath = f"{folder}/{base}.png"
        if os.path.normcase(os.path.abspath(newpath)) == os.path.normcase(os.path.abspath(f)):
            newpaths.append(newpath)
            continue
        if os.path.exists(newpath):                 # don't clobber an unrelated file
            newpaths.append(f)
            continue
        tmp = f"{folder}/__tmp_{i}_{base}.png"       # 2-step for case-only renames
        shutil.move(f, tmp)
        shutil.move(tmp, newpath)
        renames.append((os.path.basename(f), f"{base}.png"))
        newpaths.append(newpath)
    ws.cell(r, 7, value=";".join(newpaths))

# ---- 2. clear cost on boss-deck cards -----------------------------------------
NO_COST = {"Boss", "Minion", "Disaster", "Character"}
cleared = []
for r in range(2, ws.max_row + 1):
    if ws.cell(r, 2).value in NO_COST and ws.cell(r, 5).value not in ("", None):
        cleared.append((ws.cell(r, 1).value, ws.cell(r, 5).value))
        ws.cell(r, 5, value="")

# ---- 3. sanity flags ----------------------------------------------------------
# specific judgement calls (effect/name/category mismatches, likely extraction bleed)
MANUAL = {
    "Power Anomaly": "WRONG category: 'Reset all Charges to 0' is a Disaster effect, not a Character -> recategorize to Disaster.",
    "Mana Resonance": "WRONG category: mana-gain effect -> should be Mana (Enchanter class card), not Artifact.",
    "Cycle of Life": "WRONG category: support-retrieval effect -> should be Support (Druid class card), not Mana.",
    "Zither": "EFFECT mismatch: a Weapon should 'Deal 1 DMG'; current text is Bardic Inspiration's (card move) - extraction bleed.",
    "Gear Purge": "EFFECT/name mismatch: 'Gear Purge' should discard a Weapon/Artifact; current text is Critical Hit's (lowest HP 5 DMG).",
    "Critical Hit": "EFFECT/name mismatch: current text is Tactical Retreat's (Minion summon); 'Critical Hit' should be a damage hit.",
    "Unnatural Disaster": "EFFECT/name mismatch: current text is Gear Purge's (discard gear); was 'Village takes 8 DMG'.",
    "Kobold": "Should be a MINION, not Boss; effect missing - fill in.",
    "Tactical Retreat": "Effect missing - likely 'Discard lowest-HP Minion, then summon 2'.",
    "Wand": "Effect missing - Enchanter class Artifact, likely 'Deal 1 DMG', cost 0.",
    "Bully the Weak": "Verify category: marked Disaster but reads like a Boss attack.",
    "Mystic Seer's Lens": "Cost check: 6 Mana for 'Draw 2 Cards' only (no damage) looks expensive vs Arcane Tome (4) / Mystical Arcanum (5).",
}
flagged = 0
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 1).value
    cat = ws.cell(r, 2).value
    tier = ws.cell(r, 3).value
    cost = ws.cell(r, 5).value
    eff = (ws.cell(r, 6).value or "").strip()
    flags = []
    if name in MANUAL:
        flags.append(MANUAL[name])
    if not eff and "missing" not in " ".join(flags):
        flags.append("Effect missing - fill in.")
    if re.search(r"Effect:|\n", ws.cell(r, 6).value or ""):
        flags.append("Messy text (stray 'Effect:'/line breaks) - clean up.")
    if cat in ("Weapon", "Artifact") and not tier and cost not in ("", None) and isinstance(cost, int) and cost > 0:
        flags.append("Tier missing (Light/Heavy or Common/Ancient?) for a costed card.")
    if flags:
        c = ws.cell(r, SANCOL, value=" ".join(flags))
        c.fill = PatternFill("solid", fgColor="F8CBAD")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        flagged += 1

wb.save(XLSX)
print(f"renamed {len(renames)} image files")
for a, b in renames:
    print(f"   {a}  ->  {b}")
print(f"\ncleared cost on {len(cleared)} boss-deck cards: {cleared}")
print(f"\nflagged {flagged} cards in Sanity Check column")
