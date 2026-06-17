"""
Re-OCR boss-deck cards (Boss/Minion/Disaster) and VERIFY effects vs the rendered cards.
Stage 1 (cached): OCR each boss-deck cell's lines -> boss_ocr_cache.json.
Stage 2: extract effect, match each sheet card to its closest OCR effect, classify:
  confirmed (sheet effect is contained in / very close to OCR) | differs (show OCR) | not found.
Updates the Sanity column in place; prints a report.
"""
import fitz, pytesseract, io, re, json, os, difflib
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill, Alignment
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_XLSX = os.path.join(_ROOT, "data", "Cards_Data.xlsx")
_PDF = os.path.join(_ROOT, "assets", "Image Cards.pdf")
CACHE = os.path.join(_ROOT, "data", "caches", "boss_ocr_cache.json")
PAGES = list(range(26, 37))

def norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())

# ---- Stage 1: OCR cells with a 'Red Dragon' footer ----------------------------
if os.path.exists(CACHE):
    cells = json.load(open(CACHE))
else:
    doc = fitz.open(_PDF)
    cells = []
    for pidx in PAGES:
        pg = doc[pidx]
        img = Image.open(io.BytesIO(pg.get_pixmap(matrix=fitz.Matrix(3, 3)).tobytes("png")))
        cw, ch = img.width / 4, img.height / 2
        for r in range(2):
            for c in range(4):
                cell = img.crop((int(c*cw), int(r*ch), int((c+1)*cw), int((r+1)*ch)))
                lines = [l.strip() for l in pytesseract.image_to_string(cell, config="--psm 6").splitlines() if l.strip()]
                txt = " ".join(lines)
                if "Red Dragon" in txt and "Turn Overview" not in txt:
                    cells.append(lines)
        print("ocr page", pidx)
    json.dump(cells, open(CACHE, "w"))
print("boss-deck cells OCR'd:", len(cells))

# ---- effect extraction --------------------------------------------------------
def extract_effect(lines):
    ei = next((i for i, l in enumerate(lines) if re.search(r"f.?fect", l.lower())), None)
    if ei is not None:
        body = lines[ei:ei + 6]
    else:                                   # minions: no 'Effect:' header
        body = []
        for i, l in enumerate(lines):
            ll = l.lower()
            if i == 0 or "red dragon" in ll or ll.startswith("minion") \
               or re.search(r"\bhp\b", ll) or "affinity" in ll or "charge 1" in ll:
                continue
            body.append(l)
    eff = " ".join(body)
    eff = re.sub(r"^.*?f.?fect[:\s]*", "", eff, flags=re.I)
    eff = re.sub(r"\s*\S*\s*Red Dragon.*$", "", eff)      # drop footer (+ mangled word before it)
    eff = re.sub(r"[|�_]+", " ", eff).replace("0MG", "DMG")
    return re.sub(r"\s+", " ", eff).strip(" -:")

ocr_effs = [extract_effect(l) for l in cells]
ocr_effs = [e for e in ocr_effs if len(norm(e)) >= 4]

# ---- Stage 2: match each sheet boss-deck card to its closest OCR effect --------
wb = openpyxl.load_workbook(_XLSX)
ws = wb["Cards"]
green = PatternFill("solid", fgColor="C6E0B4")
amber = PatternFill("solid", fgColor="F8CBAD")
grey = PatternFill("solid", fgColor="D9D9D9")
confirmed = disc = miss = 0
diffs = []
for r in range(2, ws.max_row + 1):
    if ws.cell(r, 2).value not in ("Boss", "Minion", "Disaster"):
        continue
    name, seff = ws.cell(r, 1).value, ws.cell(r, 6).value or ""
    sn = norm(seff)
    best, br = "", 0.0
    for oe in ocr_effs:
        rt = difflib.SequenceMatcher(None, sn, norm(oe)).ratio()
        if rt > br:
            best, br = oe, rt
    contained = sn and norm(best) and (sn in norm(best) or norm(best) in sn)
    cell = ws.cell(r, 9)
    if contained or br >= 0.80:
        cell.value, cell.fill = "OCR-confirmed vs card.", green
        confirmed += 1
    elif br >= 0.5:
        cell.value, cell.fill = f'OCR could not auto-confirm (closest read, may be a neighbor card): "{best}" - verify manually', amber
        disc += 1
        diffs.append((name, seff, best))
    else:
        cell.value, cell.fill = "not located in OCR - verify manually", grey
        miss += 1
    cell.alignment = Alignment(vertical="top", wrap_text=True)
wb.save(_XLSX)

print(f"\nboss-deck: {confirmed} confirmed | {disc} differ | {miss} not found")
print("\n-- DIFFERENCES --")
for n, s, o in diffs:
    print(f"  {n}\n     sheet: {s}\n     OCR  : {o}")
