"""
Apply high-confidence corrections to Cards_Data.xlsx in place, and rewrite the Sanity
Check column to (a) mark what was corrected and (b) keep only genuinely-uncertain flags.
"""
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

_XLSX = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                     "data", "Cards_Data.xlsx")
ws = openpyxl.load_workbook(_XLSX)["Cards"]
SAN = 9

# the canonical 10-card Disaster deck (separate from Boss attacks)
DISASTERS = {"Critical Hit", "Gear Purge", "Unnatural Disaster", "Tactical Retreat",
             "Trade Block", "Power Anomaly", "Fiery Explosion", "Supply Depletion",
             "Reawakening", "Minion Frenzy"}

# category/tier/class fixes: name -> (category, tier, class-or-None to keep)
CATFIX = {
    "Mana Resonance": ("Mana", "", "Enchanter"),
    "Cycle of Life": ("Support", "", "Druid"),
    "Anger Management": ("Weapon", "", "Bard"),
    "Kobold": ("Minion", "", None),
    "Bully the Weak": ("Boss", "", None),
    "Rupture Relic": ("Artifact", "Ancient", None),
}
# effect fixes (restored from prior version / obvious from name)
EFFFIX = {
    "Gear Purge": "Every Player discards 1 Weapon or Artifact",
    "Critical Hit": "The Player with the lowest HP takes 5 DMG",
    "Unnatural Disaster": "The Village takes 8 DMG",
    "Tactical Retreat": "Discard the Minion with least HP, then summon 2 from the Discard Pile "
                        "with full HP. Summon only 1 if no Minion was discarded",
    "Zither": "Deal 1 DMG",
    "Wand": "Deal 1 DMG",
    "Kobold": "The Village takes 3 DMG. 1 Player discards 1 Card",
    "Drain the Horde": "+1 Mana. Increase by 1 for every active Minion",
}
COSTFIX = {"Wand": 0}

# minion effects + HP restored from the prior version (extraction shuffled them);
# HP goes in the Tier column (otherwise empty for minions).
MINION = {
    "Kobold":          ("4 HP",  "The Village takes 3 DMG. 1 Player discards 1 Card"),
    "Kobold Horde":    ("8 HP",  "The Village takes 5 DMG. 1 Player discards 1 Card"),
    "Kobold Archer":   ("5 HP",  "All Players take 2 DMG"),
    "Kobold Thief":    ("5 HP",  "The Player with the most Cards discards 2 Cards, or discard 3 Cards collectively"),
    "Kobold Shaman":   ("5 HP",  "1 DMG per Player collectively. 1 Player loses 1 Affinity, 1 Player loses 1 Charge"),
    "Kobold Marauder": ("5 HP",  "1 Player takes 4 DMG, or the Village takes 4 DMG"),
    "Kobold Defender": ("10 HP", "1 Player or the Village takes 1 DMG. If the Boss or another Minion would take DMG, redirect to this Minion instead"),
    "Dragon Cultist":  ("6 HP",  "1 Player takes 3 DMG. 2 Players lose 1 Charge"),
    "Dragon Lancer":   ("5 HP",  "The Player with the most Mana Cards in Hand takes 1 DMG per Card"),
    "Fire Breather":   ("5 HP",  "All Players take 2 DMG, or discard half the player count (rounded up) of Cards collectively"),
    "Wyrm":            ("8 HP",  "The Player with the highest HP takes 5 DMG"),
    "Wyrmling":        ("6 HP",  "1 Player takes 4 DMG. 1 Player discards 1 Card"),
}
RESTORED = {"Gear Purge", "Critical Hit", "Unnatural Disaster", "Tactical Retreat", "Kobold"} | set(MINION)

corrected = {}
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 1).value
    msg = []
    if name in CATFIX:
        cat, tier, kl = CATFIX[name]
        ws.cell(r, 2, cat); ws.cell(r, 3, tier)
        if kl is not None:
            ws.cell(r, 4, kl)
        msg.append(f"category->{cat}" + (f"/{tier}" if tier else ""))
    if name in DISASTERS and ws.cell(r, 2).value != "Disaster":
        ws.cell(r, 2, "Disaster"); ws.cell(r, 3, "")
        msg.append("category->Disaster")
    if name in EFFFIX:
        ws.cell(r, 6, EFFFIX[name])
        msg.append("effect restored" if name in RESTORED else "effect fixed")
    if name in COSTFIX:
        ws.cell(r, 5, COSTFIX[name]); msg.append(f"cost->{COSTFIX[name]}")
    if name in MINION:
        hp, eff = MINION[name]
        ws.cell(r, 3, hp); ws.cell(r, 6, eff)
        msg.append("HP+effect restored")
    if msg:
        corrected[name] = msg

# rewrite Sanity column
KEEP_NOTE = {
    "Mystic Seer's Lens": "Cost check: 6 Mana for 'Draw 2 Cards' only (no damage) looks high vs Arcane Tome (4) / Mystical Arcanum (5) - design call.",
}
green = PatternFill("solid", fgColor="C6E0B4")
amber = PatternFill("solid", fgColor="F8CBAD")
nfix = nflag = 0
for r in range(2, ws.max_row + 1):
    name = ws.cell(r, 1).value
    cat = ws.cell(r, 2).value
    txt, fill = "", None
    if name in corrected:
        note = "CORRECTED: " + ", ".join(corrected[name])
        if name in RESTORED:
            note += " (restored from prior version - verify exact numbers)"
        txt, fill = note, green
        nfix += 1
    elif cat == "Minion":
        txt, fill = "VERIFY effect: minion effects look shuffled in extraction - confirm vs card (recommend boss-deck re-OCR).", amber
        nflag += 1
    elif name in KEEP_NOTE:
        txt, fill = KEEP_NOTE[name], amber
        nflag += 1
    c = ws.cell(r, SAN, value=txt)
    c.fill = fill if fill else PatternFill(fill_type=None)
    c.alignment = Alignment(vertical="top", wrap_text=True)

openpyxl.Workbook  # noqa
ws.parent.save(_XLSX)
print("corrected:", nfix, "cards")
for n, m in corrected.items():
    print(f"   {n}: {', '.join(m)}")
print("still flagged:", nflag)
